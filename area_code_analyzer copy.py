import tkinter as tk
from tkinter import filedialog
from tkinter import scrolledtext
import openpyxl
import json
import pkg_resources
from PIL import Image, ImageTk
from io import BytesIO
import requests
import os
import pandas as pd

TOKEN_LOCAL = "FFL3j9HDLf"  # Token almacenado en tu programa
estados_seleccionados = []  # Lista para almacenar estados seleccionados



def convertir_csv_a_xlsx(archivo_csv):
    # Lee el archivo CSV en un DataFrame de pandas
    df = pd.read_csv(archivo_csv)

    # Pide al usuario la ubicación y el nombre del archivo XLSX resultante
    archivo_xlsx = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel Files", "*.xlsx")])

    if archivo_xlsx:
        # Guarda el DataFrame en un archivo XLSX
        df.to_excel(archivo_xlsx, index=False)
        resultado_label.config(text=f"Archivo CSV convertido a XLSX y guardado en '{archivo_xlsx}'")


def quitar_uno(numero):
    numero = str(numero)
    if len(numero) == 11 and numero.startswith("1"):
        numero = numero[1:]
    return numero

def cargar_area_codes():
    # Cargar el archivo JSON como un recurso del paquete
    try:
        area_codes_data = pkg_resources.resource_string(__name__, "area_codes.json")
        area_codes_data = area_codes_data.decode("utf-8")  # Decodificar como cadena
        area_codes_data = json.loads(area_codes_data)
    except (pkg_resources.ResourceNotFound, json.JSONDecodeError):
        # Manejo de errores si el archivo JSON no se encuentra o no se puede cargar
        area_codes_data = {}
    return area_codes_data

def procesar_archivo_excel(file_path, area_codes_data):
    wb = openpyxl.load_workbook(file_path)
    sheet = wb.active
    ubicaciones = {}

    for row in sheet.iter_rows(values_only=True):
        for numero in row:
            if isinstance(numero, str) and numero.isdigit():  # Verifica si es una cadena de dígitos
                numero = quitar_uno(numero)
                codigo_area = numero[:3]

                ubicacion = area_codes_data.get(codigo_area, None)
                if ubicacion is None:
                    for codigo_overlay, data in area_codes_data.items():
                        overlay_complex = data["Overlay complex"].split("/")
                        if codigo_area in overlay_complex:
                            ubicacion = data
                            break
                    else:
                        ubicacion = {"Location": "N/A"}

                if ubicacion["Location"] not in ubicaciones:
                    ubicaciones[ubicacion["Location"]] = []

                ubicaciones[ubicacion["Location"]].append(numero)

    return ubicaciones


def guardar_resultados(ubicaciones):
    file_path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel Files", "*.xlsx")])
    
    if file_path:
        wb_nuevo = openpyxl.Workbook()
        sheet_nuevo = wb_nuevo.active

        for idx, ubicacion in enumerate(ubicaciones):
            columna = sheet_nuevo.cell(row=1, column=idx + 1, value=ubicacion)
            for i, numero in enumerate(ubicaciones[ubicacion]):
                sheet_nuevo.cell(row=i + 2, column=idx + 1, value=numero)

        wb_nuevo.save(file_path)
        resultado_label.config(text=f"Números segmentados y guardados en '{file_path}'")

def combinar_archivos_excel():
    archivos = filedialog.askopenfilenames(filetypes=[("Excel Files", "*.xlsx")])
    
    if archivos:
        # Crear un nuevo archivo Excel para almacenar la combinación
        file_path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel Files", "*.xlsx")])
        
        if file_path:
            # Crear un nuevo libro de trabajo
            wb_nuevo = openpyxl.Workbook()
            sheet_nuevo = wb_nuevo.active
            
            # Diccionario para almacenar los números por estado
            numeros_por_estado = {}
            
            # Combinar columnas de los archivos seleccionados
            for archivo in archivos:
                area_codes_data = cargar_area_codes()
                ubicaciones = procesar_archivo_excel(archivo, area_codes_data)
                
                # Filtrar las ubicaciones por estados seleccionados
                ubicaciones_filtradas = {estado: ubicaciones[estado] for estado in estados_seleccionados if estado in ubicaciones}
                
                # Agregar los números al diccionario por estado
                for estado, numeros in ubicaciones_filtradas.items():
                    if estado not in numeros_por_estado:
                        numeros_por_estado[estado] = []
                    numeros_por_estado[estado].extend(numeros)
            
            # Agregar las ubicaciones como encabezados de columna
            for idx, estado in enumerate(numeros_por_estado):
                columna = sheet_nuevo.cell(row=1, column=idx + 1, value=estado)
                for i, numero in enumerate(numeros_por_estado[estado]):
                    sheet_nuevo.cell(row=i + 2, column=idx + 1, value=numero)
            
            # Guardar el resultado en el archivo especificado por el usuario
            wb_nuevo.save(file_path)
            resultado_label.config(text=f"Archivos combinados y guardados en '{file_path}'")



def segmentar_numeros_telefonicos():
    file_path = filedialog.askopenfilename(filetypes=[("Excel Files", "*.xlsx")])
    
    if file_path:
        area_codes_data = cargar_area_codes()
        ubicaciones = procesar_archivo_excel(file_path, area_codes_data)
        guardar_resultados(ubicaciones)


def cargar_icono():
    try:
        # Cargar el icono desde el paquete con Pillow
        icon_data = pkg_resources.resource_string(__name__, "belmont_crest.ico")
        icon = Image.open(BytesIO(icon_data))
        icon = icon.resize((64, 64))  # Ajustar el tamaño según tus necesidades
        icon_img = ImageTk.PhotoImage(icon)
        return icon_img
    except (IOError, FileNotFoundError):
        return None

def verificar_token():
    # URL donde se encuentra el token remoto
    url = "https://test.uvm.edu.ve/token.txt"
    
    try:
        # Realizar una solicitud HTTP para obtener el token remoto
        response = requests.get(url)
        if response.status_code == 200:
            token_remoto = response.text.strip()
            return token_remoto == TOKEN_LOCAL
        else:
            return False
    except requests.exceptions.RequestException:
        return False

# Función para agregar estados a la lista
def agregar_estado():
    estado = estado_entry.get()
    if estado:
        estados_seleccionados.append(estado)
        estados_listbox.insert(tk.END, estado)
        estado_entry.delete(0, tk.END)

def filtrar_numeros_por_campana(ruta_archivo, campana_seleccionada):
    try:
        # Carga el archivo Excel de Ringba
        wb_ringba = openpyxl.load_workbook(ruta_archivo)
        sheet_ringba = wb_ringba.active

        # Crear un nuevo libro de trabajo para almacenar los números filtrados
        wb_resultado = openpyxl.Workbook()
        sheet_resultado = wb_resultado.active

        # Fila actual para el libro de resultados
        fila_actual = 1

        # Encabezados de columna para el resultado
        sheet_resultado.cell(row=fila_actual, column=1, value="Campaign")
        sheet_resultado.cell(row=fila_actual, column=2, value="Caller ID")

        # Comprobar cada fila en el archivo de Ringba
        for fila in sheet_ringba.iter_rows(values_only=True):
            if len(fila) >= 4:
                campaign = fila[1]
                caller_id = fila[3]

                # Comprobar si la campaña contiene la opción seleccionada
                if campana_seleccionada.lower() in campaign.lower():
                    fila_actual += 1
                    sheet_resultado.cell(row=fila_actual, column=1, value=campaign)
                    sheet_resultado.cell(row=fila_actual, column=2, value=caller_id)

        # Pedir al usuario la ubicación y el nombre del archivo resultante
        archivo_resultado = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel Files", "*.xlsx")])

        if archivo_resultado:
            wb_resultado.save(archivo_resultado)
            resultado_label.config(text=f"Archivo filtrado y guardado en '{archivo_resultado}'")

    except Exception as e:
        resultado_label.config(text=f"Error al procesar el archivo: {str(e)}")


# Función para manejar la selección del menú desplegable
def seleccionar_campana(event):
    campana_seleccionada = campanas_var.get()
    if campana_seleccionada:
        file_path = filedialog.askopenfilename(filetypes=[("Excel Files", "*.xlsx")])
        if file_path:
            filtrar_numeros_por_campana(file_path, campana_seleccionada)

root = tk.Tk()
root.title("Area Code Analyzer")

# Cargar el icono desde el paquete
icon_img = cargar_icono()
if icon_img:
    root.tk_setPalette(background="white")
    root.iconphoto(True, icon_img)

# Configurar el tamaño de la ventana
root.geometry("500x600")  # Ajusta el tamaño según tus necesidades

# Verificar el token antes de abrir la ventana principal
if verificar_token():
    # Botón para seleccionar archivo a segmentar
    seleccionar_archivo_button = tk.Button(root, text="Seleccionar Archivo a Segmentar", command=segmentar_numeros_telefonicos)
    seleccionar_archivo_button.pack(pady=20)

    resultado_label = tk.Label(root, text="")
    resultado_label.pack()

    # Caja de texto y botón para agregar estados
    estado_label = tk.Label(root, text="Agregar Estado:")
    estado_label.pack()
    estado_entry = tk.Entry(root)
    estado_entry.pack()
    agregar_estado_button = tk.Button(root, text="Agregar", command=agregar_estado)
    agregar_estado_button.pack()

    # Lista para mostrar estados
    estados_listbox = tk.Listbox(root)
    estados_listbox.pack()

    # Botón para combinar archivos Excel
    combinar_archivos_button = tk.Button(root, text="Combinar Archivos Excel", command=combinar_archivos_excel)
    combinar_archivos_button.pack(pady=20)

    # Botón para cambiar los formatos
    convertir_csv_button = tk.Button(root, text="Convertir CSV a XLSX", command=lambda: convertir_csv_a_xlsx(filedialog.askopenfilename(filetypes=[("CSV Files", "*.csv")])))
    convertir_csv_button.pack(pady=20)

    # Dropdown menu para seleccionar una campaña
    campanas = ["ACA", "ACA Spanish", "Debt", "Debt Spanish", "Auto", "Auto Spanish", "Medicare", "Medicare Spanish"]
    campanas_var = tk.StringVar()
    campanas_dropdown = tk.OptionMenu(root, campanas_var, *campanas)
    campanas_dropdown.pack()

    # Botón para subir un archivo de Ringba y filtrar números
    filtrar_button = tk.Button(root, text="Filtrar Números Por Vertical", command=lambda: seleccionar_campana(None))
    filtrar_button.pack()

    root.mainloop()
else:
    print("Token inválido. El programa no se abrirá.")
